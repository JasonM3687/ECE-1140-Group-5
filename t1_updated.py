#!/usr/bin/env python
# -*- coding: utf-8 -*-
#
#  t1_1.py
#  
#  Copyright 2021 Brian Arrington II <brianarringtonii@macbook-pro.lan>
#  
#  This program is free software you can redistribute it and/or modify
#  it under the terms of the GNU General Public License as published by
#  the Free Software Foundation either version 2 of the License, or
#  (at your option) any later version.
#  
#  This program is distributed in the hope that it will be useful,
#  but WITHOUT ANY WARRANTY without even the implied warranty of
#  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#  GNU General Public License for more details.
#  
#  You should have received a copy of the GNU General Public License
#  along with this program if not, write to the Free Software
#  Foundation, Inc., 51 Franklin Street, Fifth Floor, Boston,
#  MA 02110-1301, USA.
#  
#  

import tkinter  
import openpyxl     
import threading           
from tkinter import *
from tkinter import ttk
from tkinter import simpledialog
from tkinter import filedialog
import datetime
from datetime import datetime
from datetime import timedelta
from PIL import ImageTk,Image
import numpy as np

class CTCOFFICE:
	
	def __init__(self,m):		
		#create Tab Window
		self.master=m
		self.tabControl = ttk.Notebook(self.master)
		tab1 = ttk.Frame(self.tabControl,width=700,height=400) 
		tab2 = ttk.Frame(self.tabControl,width=700,height=600) 
		tab3 = ttk.Frame(self.tabControl,width=700,height=400)
		tab4 = ttk.Frame(self.tabControl,width=700,height=400)  
  
		self.tabControl.add(tab1, text ='TRAIN STATUS') 
		self.tabControl.add(tab2, text ='TRACK STATUS') 
		self.tabControl.add(tab3, text ='SCHEDULE') 
		self.tabControl.add(tab4, text ='DEBUG') 
		self.tabControl.pack(expand = 1, fill ="both")
		self.end=0
		
		#--KEEP UPDATING THE TRAIN STATS--
		threading.Timer(2.0, self.GO).start()
		
		
		#rename screen title
		now = datetime.now()
		date_time = now.strftime("%H:%M %p")
		self.master.title("CTC OFFICE  -  "+date_time) 
		
		#----------------------- TAB1 ----------------------------
		self.A=["50","22","STATION 1","STATION 5","200","100"]
		self.B=["70","34","STATION 6","STATION 1","267","100"]
		self.C=["60","75","STATION 12","STATION 1","291","100"]
		
		self.Trains_current= [[0 for x in range (6)]for y in range (10)]
		
		#Set string variables for entry boxes
		self.sp=StringVar(tab1)
		self.aut= StringVar(tab1)
		self.pos=StringVar(tab1)
		self.des=StringVar(tab1)
		self.occ=StringVar(tab1)
		self.tp=StringVar(tab1)
		self.tpred=StringVar(tab1)
		
		#Create Entry boxes
		self.E=  ttk.Entry(tab1,width=35,textvariable=self.sp,foreground='blue').place(x=255,y=195)
		self.E1= ttk.Entry(tab1,width=35,textvariable=self.aut,foreground='blue').place(x=255,y=250)
		self.E2= ttk.Entry(tab1,width=35,state=DISABLED,textvariable=self.pos,foreground='blue').place(x=255,y=305)
		self.E3= ttk.Entry(tab1,width=35,state=DISABLED,textvariable=self.des,foreground='blue').place(x=255,y=360)
		#self.E4= ttk.Entry(tab1,width=35,state=DISABLED,textvariable=self.occ,foreground='blue').place(x=255,y=415)
		self.E5= ttk.Entry(tab1,width=15,state=DISABLED,textvariable=self.tp,foreground='green').place(x=240,y=495)
		self.E5= ttk.Entry(tab1,width=15,state=DISABLED,textvariable=self.tpred,foreground='red').place(x=240,y=445)
		
		#Create Option Menu
		self.Options = ["--Select--","TRAIN 1","TRAIN 2","TRAIN 3","TRAIN 4","TRAIN 5","TRAIN 6","TRAIN 7","TRAIN 8","TRAIN 9","TRAIN 10"]
		self.tkvarq = StringVar(tab1) 
		self.tkvarq.set(self.Options[0])
		self.o=ttk.OptionMenu(tab1,  self.tkvarq, *self.Options)
		self.o.config(width=33)
		self.o.place(x=246,y=130)
		
		#Create Labels + Logout Button
		self.B1 = ttk.Button(tab1, text="Logout", width=8,command=self.OUT).place(x=580,y=550)
		self.Bnew = ttk.Button(tab1, text="Update", width=8).place(x=480,y=495)
		self.B2 = ttk.Button(tab1, text="GO ->", width=4,command=self.GO).place(x=586,y=130)
		self.B3 = ttk.Label(tab1, text="TRAIN STATUS",font=('Helvetica',19,'bold'),foreground='blue').place(x=35,y=65)
		self.B4 = ttk.Label(tab1, text="SPEED (Mph)").place(x=135,y=195)
		self.B5 = ttk.Label(tab1, text="AUTHORITY").place(x=135,y=250)
		self.B6 = ttk.Label(tab1, text="POSITION").place(x=135,y=305)
		self.B7 = ttk.Label(tab1, text="DESTINATION").place(x=135,y=360)
		#self.B8 = ttk.Label(tab1, text="OCCUPANCY").place(x=135,y=415)
		self.B9 = ttk.Label(tab1, text="THROUGHPUT GREEN(Tickets/Hour)",font=('Helvetica',12,'bold')).place(x=35,y=500)
		self.Thru2 = ttk.Label(tab1, text="THROUGHPUT RED(Tickets/Hour)",font=('Helvetica',12,'bold')).place(x=35,y=450)
		
		#----------------------- TAB 2 ----------------------------
		
		#Create Variables 
		self.switches= list()
		self.MM=IntVar(tab2)
		self.closure= StringVar(tab2)
		self.im = Image.open("Track.png")
		self.im= self.im.resize((275,350),Image.ANTIALIAS)
		self.img = ImageTk.PhotoImage(self.im)
		
		#Create Labels + ListBoxes + Buttons
		self.V1 =ttk.Label(tab2, text="TRACK STATUS",font=('Helvetica',19,'bold'),foreground='blue').place(x=35,y=65)
		self.V2 =ttk.Label(tab2, text="TRACK SECTION").place(x=51,y=180)
		self.V3 =ttk.Label(tab2, text="SWITCHES").place(x=232,y=180)
		self.l6= Listbox(tab2,height=12,width=17)
		self.l6.place(x=35,y=215)
		self.l7= Listbox(tab2,height=12,width=10)
		self.l7.place(x=220,y=215)
		self.l8= ttk.Label(tab2,image=self.img)
		self.l8.place(x=370,y=100)
		self.E17=Entry(tab2,width=15,relief=FLAT,textvariable=self.closure)
		self.E17.place(x=35,y=435)
		self.V4= ttk.Button(tab2, text="Upload", width=8,command=self.PIC_UPDATE).place(x=470,y=475)
		self.V5 =ttk.Button(tab2, text="Logout", width=8,command=self.OUT).place(x=580,y=550)
		self.plus = ttk.Button(tab2, text="+", width=3,state=DISABLED,command=self.AddC)
		self.plus.place(x=35,y=475)
		self.minus= ttk.Button(tab2, text="-", width=3,state=DISABLED,command=self.CancelC)
		self.minus.place(x=98,y=475)
		self.togg= ttk.Button(tab2, text="Toggle",state=DISABLED,command=self.TOGGLE, width=6)
		self.togg.place(x=225,y=475)
		self.r2= Checkbutton(tab2, text="MAINTINENCE MODE",bg='grey',relief=FLAT,variable=self.MM,command=self.Mode)
		self.r2.place(x=35,y=120)
		

		#----------------------- TAB 3 ----------------------------
		
		#Create Variables V6
		
		self.T= StringVar(tab3)
		self.D= StringVar(tab3) 
		self.ti=StringVar(tab3) 
		self.s=StringVar(tab3) 
		self.a=StringVar(tab3) 
		self.autodisp= IntVar(tab3)
		self.routes=[]
		self.sections=list()
		self.track_blocks=list()
		self.line_used=list()
		self.CurrentTravel= []
		self.CurrentTravel_line_speed=[]
		self.SentRouteBlocks=list()
		self.SentRouteSections=list()
		self.SentRouteLines=list()
		self.SentSuggestedSpeed= []
		self.SentSuggestedAuth=[]
		#Create Entries + Listboxes + Buttons
		
		self.E6 =  Entry(tab3,width=6,textvariable=self.T,relief=FLAT).place(x=37,y=440)
		self.E7 =  Entry(tab3,width=30,textvariable=self.D,relief=FLAT).place(x=119,y=440)
		self.E8 =  Entry(tab3,width=6,textvariable=self.ti,relief=FLAT).place(x=415,y=440)
		self.E9 =  Entry(tab3,width=6,textvariable=self.s,relief=FLAT).place(x=495,y=440)
		self.E10 = Entry(tab3,width=6,textvariable=self.a,relief=FLAT).place(x=576,y=440)
		
		self.V6 = ttk.Button(tab3, text="Upload", width=8,command=self.UPLOAD_SCHEDULE).place(x=180,y=495)
		self.V7 = ttk.Button(tab3, text="+", width=3,command=self.Add).place(x=30,y=495)
		self.V8 = ttk.Button(tab3, text="-", width=3,command= self.Remove).place(x=100,y=495)
		self.disp= ttk.Button(tab3, text="Dispatch",command=self.GetDispatchInfo, width=12)
		self.disp.place(x=245,y=550)		
		
		
		self.V9 = ttk.Button(tab3, text="Logout",width=8,command=self.OUT).place(x=580,y=550)
		self.V10 =ttk.Label(tab3, text="SCHEDULE",font=('Helvetica',19,'bold'),foreground='blue').place(x=35,y=60)
		self.V11 =ttk.Label(tab3, text="TRAIN").place(x=45,y=180)
		self.V12 =ttk.Label(tab3, text="      ROUTE").place(x=195,y=180)
		self.V13 =ttk.Label(tab3, text="   ARRIVE").place(x=410,y=180)
		self.V14 =ttk.Label(tab3, text="SPEED(Km/Hr)").place(x=488,y=180)
		self.V15 =ttk.Label(tab3, text="AUTH.(m)").place(x=580,y=180)


		self.l= Listbox(tab3,height=12,width= 8)
		self.l.place(x=35,y=215)


		self.l2= Listbox(tab3,height=12,width=32)
		self.l2.place(x=115.5,y=215)

		self.l3= Listbox(tab3,height=12,width=8)
		self.l3.place(x=413,y=215)

		self.l4= Listbox(tab3,height=12,width=8)
		self.l4.place(x=493,y=215)

		self.l5= Listbox(tab3,height=12,width=8)
		self.l5.place(x=573,y=215)

		self.r= Checkbutton(tab3, text="AUTO-DISPATCH",bg='grey',relief=FLAT,variable=self.autodisp,command= self.Auto)
		self.r.place(x=35,y=120)
		
		
		#----------------------- TAB 4 ---------------------------- V16
		
		#CREATE GUI FOR DEBUGGING
		self.V16 = ttk.Label(tab4, text="DEBUG",font=('Helvetica',19,'bold')).place(x=35,y=60)
		self.V17 = ttk.Label(tab4, text="TRACK DEBUG",font=('Helvetica',16,'bold')).place(x=35,y=375)

		self.E11=ttk.Entry(tab4,width=15)
		self.E11.place(x=85,y=100)
		self.E12=ttk.Entry(tab4,width=15)
		self.E12.place(x=85,y=140)
		self.E13=ttk.Entry(tab4,width=15)
		self.E13.place(x=85,y=180)
		self.E14=ttk.Entry(tab4,width=15)
		self.E14.place(x=85,y=220)
		self.E15=ttk.Entry(tab4,width=15)
		self.E15.place(x=85,y=260)
		self.E16=ttk.Entry(tab4,width=15)
		self.E16.place(x=85,y=300)

		self.V18 = ttk.Label(tab4, text="SPEED").place(x=35,y=100)
		self.V19 = ttk.Label(tab4, text="AUTH").place(x=35,y=140)
		self.V20 = ttk.Label(tab4, text="POS").place(x=35,y=180)
		self.V21 = ttk.Label(tab4, text="DES").place(x=35,y=220)
		self.V22 = ttk.Label(tab4, text="OCC").place(x=35,y=260)
		self.V23 = ttk.Label(tab4, text="THPT").place(x=35,y=300)

		self.Options2 = ["--Select--","TRAIN A","TRAIN B","TRAIN C"]
		self.tkvarq2 = StringVar(tab4) 
		self.tkvarq2.set(self.Options2[0])
		self.o2=ttk.OptionMenu(tab4,  self.tkvarq2, *self.Options2)
		self.o2.config(width=20)
		self.o2.place(x=85,y=340)

		self.V24 = ttk.Button(tab4, text="GO ->",command=self.DEBUGGER).place(x=310,y=340)
		self.V25 = ttk.Button(tab4, text="GO ->",command=self.UPDATE).place(x=35,y=500)


		self.V26 =ttk.Label(tab4, text="CLOSED SECTIONS: ").place(x=35,y=415)
		self.V27 = ttk.Label(tab4, text="SWITCH STATES: ").place(x=35,y=460)

		self.CS =ttk.Label(tab4, text="")
		self.CS.place(x=200,y=415)
		self.SS=ttk.Label(tab4, text="")
		self.SS.place(x=200,y=460)
		
		#self.master.mainloop()
		
	#TAB 1 FUNCTIONS
	def GO(self):
		
	
		
		delete_later=list() 
		
		row_indexer=["Train 1","Train 2","Train 3","Train 4","Train 5","Train 6","Train 7","Train 8","Train 9","Train 10"]
		for i in range(0,len(self.CurrentTravel)):
			#if length of current travel is 0 set array to 0 and break
			temp=self.CurrentTravel[i]
			write_row=0
			
			indexing=i*2
			temp2=self.CurrentTravel_line_speed[indexing]
			temp3=self.CurrentTravel_line_speed[indexing+1]
			
			#Determine if train is on red line or green line load the correct Occupancy and Authority 
			if(temp2==1):
				print("GREEN LINE FUNCTION")
				BlockOccupancy=[0,0,0,0,1,0,0,0,0,0] #TEST VALUES
				print("GREEN LINE AUTHORITY")
				BlockAuthority=[0,7,6,5,4,3,2,1,0,0] #TEST VALUES
				print("CALL THROUGHPUT FUNCTIONS")
				Throughput=[0,0,0,0] #TEST VALUES
				Throughput_red=[0,0,0,100] #TEST VALUES
			else:
				print("RED LINE FUNCTION")
				BlockOccupancy=[0,0,0,0,0,0,0,0,0,0] #TEST VALUES
				print("RED LINE AUTHORITY")
				BlockAuthority=[0,0,0,0,0,0,0,0,0,0] #TEST VALUES
				print("CALL THROUGHPUT FUNCTIONS")
				Throughput=[0,0,0,0] #TEST VALUES
				Throughput_red=[0,70,0,0] #TEST VALUES
			
			#get Train ID
			for y in range(0,len(row_indexer)):
				if(temp[len(temp)-1] == row_indexer[i]):
					write_row=i
			#get the route and store the value of each block as an int
			path=[]
			for m in range(0,len(temp)-1):
				path.append(int(temp[m]))
			#Gon through the track look for occupancy, if found see if the occupancy is along the route of the current train
			for z in range(0,len(BlockOccupancy)):
				if(BlockOccupancy[z]==1 and z+1 in path):
					if(BlockAuthority[z]== 0):
						self.Trains_current[write_row][0]= 0
					else:
						self.Trains_current[write_row][0]= temp3
					self.Trains_current[write_row][1]= BlockAuthority[z]
					self.Trains_current[write_row][2]= z+1 #BLOCK #
					self.Trains_current[write_row][3]= temp[len(temp)-2] #destination block
					self.Trains_current[write_row][4]= str(Throughput)
					self.Trains_current[write_row][5]= str(Throughput_red)		
				if(z+1==int(temp[len(temp)-2])):
					delete_later.append(i)
		
		#remove items that have reached their destination	
		for q in range(0,len(delete_later)):
				del self.CurrentTravel[delete_later[q]]
			
		self.Disp()
	
		#repeat every second add a condition so it checks if the text box still matches	the text file	
		if(self.end==0):
			threading.Timer(2.0, self.GO).start()
		return		
	def Disp(self):
		for i in range (0,len(self.Options)):
			if(self.tkvarq.get()==self.Options[i] and self.tkvarq.get!="--Select--"):
				self.sp.set(self.Trains_current[i-1][0])
				self.aut.set(self.Trains_current[i-1][1])
				self.pos.set(self.Trains_current[i-1][2])
				self.des.set(self.Trains_current[i-1][3])
				self.tp.set(self.Trains_current[i-1][4])
				self.tpred.set(self.Trains_current[i-1][5])	
			
		
	#TAB 2 FUNCTIONS
	def Mode(self):
		if(self.MM.get()==0):
			self.plus.config(state=DISABLED)
			self.minus.config(state=DISABLED)
			self.togg.config(state=DISABLED)
		else:
			self.plus.config(state=NORMAL)
			self.minus.config(state=NORMAL)
			self.togg.config(state=NORMAL)
		return
	def AddC(self):
		t=self.l6.get(0,"end")
		test=self.list_to_string(self.closure)
		if((self.closure.get() in t) or test==""):
			return
		else:
			self.l6.insert(self.l6.size()+1,self.closure.get())
			self.closure.set("")
		return	
	def CancelC(self):
		x=self.l6.curselection()
		if(x != tuple()):
			self.l6.delete(x)
		return 
	def TOGGLE(self):
		
		x=self.l7.curselection()
		if(x !=tuple() and len(self.switches)>0):
			if(self.switches[x[0]] == "OFF"):
				self.switches[x[0]] ="ON"
				
			else:
				self.switches[x[0]]="OFF"
				
		import os
		title= "TRACK UPDATE"
		x=self.l7.curselection()
		if(len(self.switches)>0):
			message="SWITCH IS "+str(self.switches[x[0]])
		command = f'''
		osascript -e 'display notification "{message}" with title "{title}"'
		'''
		os.system(command)
		return
	def list_to_string(self,s):
		z=list(s.get())
		for b in range(0,len(z)):
			if(z[b]==' '):
				z[b]=''
		ingo= ''.join(map(str,z))
		return ingo
	def PIC_UPDATE(self):
		loc= filedialog.askopenfilename()
		self.im2 = Image.open(loc)
		self.im2= self.im2.resize((275,350),Image.ANTIALIAS)
		self.img2 = ImageTk.PhotoImage(self.im2)
		self.l8.config(image=self.img2)


		
	#TAB 3 FUNCTIONS
	def Remove(self):
		x=self.l.curselection()
		del self.routes[0]
		if(x !=tuple()):
			self.l.delete(x)
			self.l2.delete(x)
			self.l3.delete(x)
			self.l4.delete(x)
			self.l5.delete(x)
		return
	def Add(self):
		check=[self.T,self.D,self.ti,self.a,self.s]
		for u in range(0,len(check)):
			temp=self.list_to_string(check[u])
		if(temp==""):
			return
		x=self.l.curselection()
		if(x==tuple()):
			x=self.l.size()+1
		self.l.insert(x,self.T.get())
		self.l2.insert(x,self.D.get())
		self.l3.insert(x,self.ti.get())
		self.l4.insert(x,self.s.get())
		self.l5.insert(x,self.a.get())
		self.T.set("")
		self.D.set("")
		self.ti.set("")
		self.a.set("")
		self.s.set("")
		return	
	def Auto(self):
		if(self.autodisp.get()==1):
			self.disp.config(state=DISABLED)
		else:
			self.disp.config(state=NORMAL)
		return
	def UPLOAD_SCHEDULE(self):
		
		#loc= "schedule.xlsx"
		loc= filedialog.askopenfilename()
		wb = openpyxl.load_workbook(loc)
		sc= wb.active
		rows= sc.max_row
		cols= sc.max_column
		recieved_Time=list()
		recieved_Route=list()
		recieved_AUTH=list()
		Train_Num=list()
		S_perBlock= list()
		S_real=list()
		Block_route= list()
		c=0
		T=list()
		holder=list()
		D=list()
		far=list()
		count=0
		meters_sep=0
		for i in range (2,rows):
			if(sc.cell(i,32).value is not None and sc.cell(i,7).value is not None):
				
				S_perBlock.append(sc.cell(i,6).value)
				Block_route.append(sc.cell(i,3).value)
				self.track_blocks.append(sc.cell(i,3).value);
				if(count>0):
					far.append(count+1)
					meters_sep=0
					count=0
					slowest=S_perBlock[0]
					for t in range (0,len(S_perBlock)):
						if(slowest>S_perBlock[t]):
							slowest=S_perBlock[t]
					S_real.append(slowest)
					S_perBlock.clear()
					
					self.routes.append(np.array(Block_route))
					
					Block_route.clear()
					
					
				holder.append(sc.cell(i,32).value)
				temp2=str(sc.cell(i,7).value)
				self.line_used.append(sc.cell(i,1).value)
				self.sections.append(sc.cell(i,2).value)
				D.append(temp2.strip())
				Train_Num.append("Train 1")
				
			if(sc.cell(i,32).value is  None):
					count=count+1
					if(sc.cell(i,4).value is not None):
						meters_sep= meters_sep+int(sc.cell(i,4).value)
						Block_route.append(sc.cell(i,3).value)
						self.track_blocks.append(sc.cell(i,3).value);
					#print (sc.cell(i,4).value)
					self.sections.append(sc.cell(i,2).value)
					self.line_used.append(sc.cell(i,1).value)
					S_perBlock.append(sc.cell(i,6).value)
					
		
		
		now=datetime.now()
		del holder[0]
		del Train_Num[0]
		initial=0
		stop=len(D)-1
		holdval= S_real
		
#		for i in range (2,11):
#			if(i>2):
#				initial=checkpoint
#			checkpoint= len(holder)
#			S_real=S_real+holdval
			#calculate all of the time value
#			for z in range(initial,checkpoint):
#				Train_Num.append("Train "+str(i))
#				holder.append((datetime.combine(now.date(),holder[z])+timedelta(minutes=sc.cell(1,43).value.minute, hours=sc.cell(1,43).value.hour)).time())
	
	#ONLY UPLOAD SCHEDULE FOR 1 TRAIN
		for i in range (0,len(D)-1):
			
			recieved_Route.append(str(D[i])+" -> "+str(D[i+1]))
		recieved_AUTH= recieved_AUTH+far
#		for i in range (0,10):
#			for i in range (0,len(D)-1):
#				recieved_Route.append(str(D[i])+" -> "+str(D[i+1]))	
#			recieved_AUTH= recieved_AUTH+far
		
		#repeat the block # routes for all trains
#		rep_len= len(self.routes)	
		
#		for i in range (0,9):
#			for z in range(0,rep_len):
#				self.routes.append(self.routes[z])
					
		
		#SORTING BY TIME
		for i in range (0,len(holder)):
			for z in range (0,len(holder)-1):
				if(holder[z]>holder[z+1]):
					switch=holder[z+1]
					holder[z+1]=holder[z]
					holder[z]=switch
				
					switch=recieved_Route[z+1]
					recieved_Route[z+1]=recieved_Route[z]
					recieved_Route[z]= switch
					
					switch=recieved_AUTH[z+1]
					recieved_AUTH[z+1]=recieved_AUTH[z]
					recieved_AUTH[z]= switch
					
					switch= Train_Num[z+1]
					Train_Num[z+1]=Train_Num[z]
					Train_Num[z]=switch
					
					switch= S_real[z+1]
					S_real[z+1]= S_real[z]
					S_real[z]=switch	
				
					switch= self.routes[z+1]
					self.routes[z+1]=self.routes[z]
					self.routes[z]= switch				
		
		for i in range (0,len(holder)):
			temp=str(holder[i])
			T.append(temp[-8:-3])
		
		#UPDATE SCHEDULE LIST BOXES
		
		self.l.delete(0,END)
		self.l2.delete(0,END)
		self.l3.delete(0,END)
		self.l4.delete(0,END)
		self.l5.delete(0,END)
		
		for i in range (0,len(Train_Num)):
			self.l.insert(i,Train_Num[i])
			self.l2.insert(i,recieved_Route[i])
			self.l3.insert(i,T[i])
			self.l5.insert(i,recieved_AUTH[i])
			self.l4.insert(i,S_real[i])
		
		#UPDATE TRACK MODEL
		
		Track_Switches= list()
		
		for i in range (2,rows):
			if("SWITCH" in str(sc.cell(i,7).value).upper()):
				Track_Switches.append(str(sc.cell(i,2).value)+str(sc.cell(i,3).value)+"-"+str(sc.cell(i,1).value))
		
		self.l7.delete(0,END)
		self.l6.delete(0,END)
		self.switches.clear()
		for i in range (0,len(Track_Switches)):
			self.l7.insert(i,Track_Switches[i])
			self.switches.insert(i,"OFF")
		#USER FEEDBACK
		import os
		
		title= "SYSTEM UPDATE"
		x=self.l7.curselection()
		message="SCHEDULE AND TRACK MODEL SUCCESSFULLY UPDATED"
		command = f'''
		osascript -e 'display notification "{message}" with title "{title}"'
		'''
		os.system(command)	
			
	def GetDispatchInfo(self):
		#for sections A=0, B=1, ......
		#for lines Green=1, red =0
		#L1= route block #, L2= section of track, L3=line
		#make all lists + speed+ auth as inputs to the function
		
		self.SentRouteBlocks.clear()
		self.SentRouteSections.clear()
		self.SentRouteLines.clear()
		self.SentSuggestedSpeed.clear()
		self.SentSuggestedAuth.clear()
		
		if(len(self.routes)>0):
			temp = self.routes[0]
			for i in range(0,len(self.routes[0])):
				self.SentRouteBlocks.append(format(temp[i],"08b"))
				calc= self.track_blocks.index(temp[i])
				#print("CALC IS : "+str(calc));
				self.SentRouteSections.append(format((ord(self.sections[calc]))-65,"08b"))
				if(str(self.line_used[calc])=="Green"):
					self.SentRouteLines.append(1)
				else:
					self.SentRouteLines.append(0)
			
			#Make sure to verify with all routes
			for i in range(0,self.l5.get(0)):
				self.SentSuggestedSpeed.append(format(self.l4.get(0),"08b"))
				self.SentSuggestedAuth.append(format(self.l5.get(0)-i,"08b"))
			self.SentSuggestedAuth.append(format(0,"08b"))
			self.SentSuggestedSpeed.append(format(self.l4.get(0),"08b"))
			
			'''print(self.SentRouteSections)
			print(self.SentRouteLines)
			print(self.SentSuggestedSpeed)
			print(self.SentSuggestedAuth)'''
			temp_list= self.routes[0]
			temp_list= np.append(temp_list,self.l.get(0))
			self.CurrentTravel.append(temp_list)
			self.CurrentTravel_line_speed.append(self.SentRouteLines[0])
			self.CurrentTravel_line_speed.append(self.l4.get(0))
			self.Remove()
			
		
	#TAB 4 FUNCTIONS
	def DEBUGGER(self):
		if(self.tkvarq2.get()=="TRAIN A"):
			self.A[0]=self.E11.get()
			self.A[1]=self.E12.get()
			self.A[2]=self.E13.get()
			self.A[3]=self.E14.get()
			self.A[4]=self.E15.get()
			self.A[5]=self.E16.get()
		if(self.tkvarq2.get()=="TRAIN B"):
			self.B[0]=self.E11.get()
			self.B[1]=self.E12.get()
			self.B[2]=self.E13.get()
			self.B[3]=self.E14.get()
			self.B[4]=self.E15.get()
			self.B[5]=self.E16.get()
		if(self.tkvarq2.get()=="TRAIN C"):
			self.C[0]=self.E11.get()
			self.C[1]=self.E12.get()
			self.C[2]=self.E13.get()
			self.C[3]=self.E14.get()
			self.C[4]=self.E15.get()
			self.C[5]=self.E16.get()
		return

	def UPDATE(self):
		self.CS.config(text=self.l6.get(0,"end"))
		self.SS.config(text=self.switches)
		return
	def OUT(self):
		self.end=1
		for item in self.tabControl.winfo_children():
			item.destroy()
		return 
 



#Make all modules take in a master value for their display window,  
#in the main make a root TK() then create windows, send the windows to the modules 
def main(args):
	master=Tk()
	c= CTCOFFICE(master)
	master.mainloop()
	

if __name__ == '__main__':
    import sys
    sys.exit(main(sys.argv))
